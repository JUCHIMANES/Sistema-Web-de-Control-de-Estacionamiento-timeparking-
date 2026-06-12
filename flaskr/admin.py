from math import ceil
from datetime import datetime, timedelta
from flask import (
    Blueprint, flash, g, redirect, render_template, make_response, request, url_for
)

from flaskr.db import get_db
from .utils import mexico_tz
from . import auth
from .utils import parse_reservation
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

bp = Blueprint('admin', __name__, url_prefix='/admin')

@bp.route("/")
@auth.only_admin
def index():
    db = get_db()
    messages = db.execute(
        'SELECT * FROM message WHERE readed = 0'
    ).fetchall()

    for message in messages:
        flash(message["message"], "success")
        db.execute(
            'UPDATE message SET readed = 1 WHERE id = ?',
            (message["id"],)
        )
    db.commit()

    # --- FILTRAR ESPACIOS DISPONIBLES PARA EL MODAL ---
    occupied_app = db.execute(
        'SELECT space_id FROM reservation WHERE status IN ("reserved", "waiting", "confirm-payment") OR (entry_datetime IS NOT NULL AND exit_datetime IS NULL)'
    ).fetchall()
    
    occupied_assisted = db.execute(
        'SELECT space_id FROM assisted_reservation WHERE status = "active"'
    ).fetchall()
    
    occupied_ids = [r['space_id'] for r in occupied_app] + [r['space_id'] for r in occupied_assisted]

    if occupied_ids:
        placeholders = ','.join(['?'] * len(occupied_ids))
        all_spaces = db.execute(
            f'SELECT * FROM space WHERE id NOT IN ({placeholders})', occupied_ids
        ).fetchall()
    else:
        all_spaces = db.execute('SELECT * FROM space').fetchall()

    current_price = db.execute(
        'SELECT * FROM price WHERE id = 1'
    ).fetchone()["price"]

    # Buscamos los dos precios en la base de datos
    price_data = db.execute('SELECT * FROM price WHERE id = 1').fetchone()
    current_price = price_data["price"]
    daily_price = price_data["daily_price"] # Agregamos esto
    
    return render_template(
        "admin/index.html", 
        current_price=current_price, 
        daily_price=daily_price, # Lo pasamos a la plantilla
        all_spaces=all_spaces
    )

@bp.route('/register-assisted', methods=['POST'])
@auth.only_admin
def register_assisted():
    db = get_db()
    
    # Captura de datos del formulario
    first_name = request.form['first_name']
    last_name = request.form['last_name']
    space_id = int(request.form['space_id'])
    entry_str = request.form['entry_datetime']

    # Convertir fecha a timestamp
    try:
        entry_naive = datetime.strptime(entry_str, "%Y-%m-%dT%H:%M")
        entry_local = mexico_tz.localize(entry_naive)
        entry_timestamp = int(entry_local.timestamp())
        
        # Validar si el espacio está ocupado en reservaciones normales
        conflict_reg = db.execute(
            'SELECT id FROM reservation WHERE space_id = ? AND status IN ("reserved", "waiting", "active")',
            (space_id,)
        ).fetchone()

        # Validar si el espacio está ocupado en reservaciones asistidas
        conflict_ast = db.execute(
            'SELECT id FROM assisted_reservation WHERE space_id = ? AND status = "active"',
            (space_id,)
        ).fetchone()

        if conflict_reg or conflict_ast:
            flash(f"El Espacio {space_id} ya está ocupado o reservado.", "error")
            return redirect(url_for("admin.index"))

        # Crear código único
        code = f"AST-{entry_timestamp}"
        current_price = db.execute('SELECT price FROM price WHERE id = 1').fetchone()["price"]

        db.execute(
            '''
            INSERT INTO assisted_reservation 
            (first_name, last_name, space_id, entry_datetime, status, cost, code)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ''',
            (first_name, last_name, space_id, entry_timestamp, "active", current_price, code)
        )
        db.commit()
        flash(f"Registro exitoso para {first_name}. Código: {code}", "success")

    except Exception as e:
        print("Error en registro asistido:", e)
        flash("Error al procesar el registro. Revisa los datos.", "error")

    return redirect(url_for("admin.index"))

@bp.route("/cancel-reservation/<reservation_id>")
@auth.only_admin
def cancel_reservation(reservation_id):
    db = get_db()
    error = None

    if not reservation_id:
        error = "El código de reservación es requerido."

    if error:
        flash(error, "error")
        return redirect(url_for("admin.history"))

    try:
        reservation = db.execute(
            'SELECT * FROM reservation WHERE id = ?', (reservation_id,)
        ).fetchone()
        current_price = db.execute(
            'SELECT * FROM price WHERE id = 1'
        ).fetchone()["price"]
        db.execute(
            'UPDATE reservation SET status = "canceled", cost = ? WHERE id = ?',
            (current_price, reservation_id)
        )
        db.commit()
        flash("Reservación cancelada correctamente", "success")
    except Exception as e:
        print("Ocurrió un error:", e)
        flash("Error al cancelar la reservación", "error")

    return redirect(url_for("admin.history"))

@bp.route("/confirm-entry", methods=["POST"])
@auth.only_admin
def confirm_entry():
    db = get_db()
    reservation_code = request.form.get("reservation_code", "").strip()
    error = None

    if not reservation_code:
        error = "El código de reservación es requerido."

    if error:
        flash(error, "error")
        return redirect(url_for("admin.index"))

    try:
        reservation = db.execute(
            'SELECT * FROM reservation WHERE code = ?', (reservation_code,)
        ).fetchone()

        if not reservation:
            flash("Reservación no encontrada", "error")
            return redirect(url_for("admin.index"))
    
        if reservation["entry_datetime"] is not None:
            flash("La reservación ya tiene una entrada registrada", "error")
            return redirect(url_for("admin.index"))

        db.execute(
            'UPDATE reservation SET entry_datetime = ? WHERE code = ?',
            (int(datetime.now(mexico_tz).timestamp()), reservation_code)
        )
        db.commit()
        flash("Entrada confirmada correctamente", "success")
    except Exception as e:
        print("Ocurrió un error:", e)
        flash("Error al confirmar la entrada", "error")

    return redirect(url_for("admin.index"))

@bp.route("/confirm-exit", methods=["POST"])
@auth.only_admin
def confirm_exit():
    db = get_db()
    reservation_code = request.form.get("reservation_code", "").strip()
    error = None
    exit_datetime = int(datetime.now(mexico_tz).timestamp())

    if not reservation_code:
        error = "El código de reservación es requerido."

    if error:
        flash(error, "error")
        return redirect(url_for("admin.index"))

    try:
        reservation = db.execute(
            'SELECT * FROM reservation WHERE code = ?', (reservation_code,)
        ).fetchone()

        if not reservation:
            flash("Reservación no encontrada", "error")
            return redirect(url_for("admin.index"))
    
        if reservation["exit_datetime"] is not None:
            flash("La reservación ya tiene una salida registrada", "error")
            return redirect(url_for("admin.index"))

        db.execute(
            'UPDATE reservation SET exit_datetime = ?, status = ? WHERE code = ?',
            (exit_datetime, "completed", reservation_code)
        )

        db.commit()
        flash(f"Salida confirmada correctamente.", "success")
        
    except:
        flash("Error al confirmar la salida", "error")

    return redirect(url_for("admin.index"))

@bp.route("/history")
@auth.only_admin
def history():
    db = get_db()
    parsed_reservations = []
    date_str = request.args.get("date")
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, "%Y-%m-%d")
            selected_date = mexico_tz.localize(selected_date)
        except ValueError:
            flash("Fecha inválida", "error")
            return redirect(url_for("admin.history"))
    else:
        selected_date = datetime.now(mexico_tz).replace(hour=0, minute=0, second=0, microsecond=0)

    start_timestamp = int(selected_date.timestamp())
    end_timestamp = int((selected_date + timedelta(days=1)).timestamp())

    try:
        reservations = db.execute(
            'SELECT * FROM reservation WHERE reservation_datetime BETWEEN ? AND ?',
            (start_timestamp, end_timestamp)
        ).fetchall()

        for reservation in reservations:
            parsed_reservation = parse_reservation(reservation)
            user = db.execute(
                'SELECT * FROM user WHERE id = ?', (parsed_reservation["user"],)
            ).fetchone()
            parsed_reservation["user"] = f"{user['name']} {user['last_name']}"
            parsed_reservations.append(parsed_reservation)

    except:
        flash("Error al obtener las reservaciones", "error")

    return render_template("admin/history.html", reservations=parsed_reservations, selected_date=selected_date.strftime("%Y-%m-%d"))

@bp.route("/assisted-history")
@auth.only_admin
def assisted_history():
    db = get_db()
    date_str = request.args.get("date")
    
    # Manejo de la fecha
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, "%Y-%m-%d")
            selected_date = mexico_tz.localize(selected_date)
        except ValueError:
            flash("Fecha inválida", "error")
            return redirect(url_for("admin.assisted_history"))
    else:
        selected_date = datetime.now(mexico_tz).replace(hour=0, minute=0, second=0, microsecond=0)

    start_timestamp = int(selected_date.timestamp())
    end_timestamp = int((selected_date + timedelta(days=1)).timestamp())

    try:
        # Obtener reservaciones asistidas del día
        raw_reservations = db.execute(
            '''SELECT * FROM assisted_reservation 
            WHERE (type = 'hour' OR type IS NULL) 
            AND entry_datetime BETWEEN ? AND ? 
            ORDER BY entry_datetime DESC''',
            (start_timestamp, end_timestamp)
        ).fetchall()
        
        # Procesar los datos para la plantilla
        reservations = []
        for r in raw_reservations:
            res = dict(r)
            
            # 1. Fecha completa (Ej: 2026-03-17)
            res["date_full"] = datetime.fromtimestamp(res["entry_datetime"], mexico_tz).strftime("%Y-%m-%d")
            
            # 2. Solo la hora de entrada (Ej: 02:30 PM)
            res["entry_time"] = datetime.fromtimestamp(res["entry_datetime"], mexico_tz).strftime("%I:%M %p")
            
            # 3. Solo la hora de salida (o Pendiente)
            if res["exit_datetime"]:
                res["exit_str"] = datetime.fromtimestamp(res["exit_datetime"], mexico_tz).strftime("%I:%M %p")
            else:
                res["exit_str"] = "Pendiente"

            res["full_name"] = f"{res['first_name']} {res['last_name']}"
            reservations.append(res)

        # Obtener precio actual para cálculos en el frontend
        current_price = db.execute('SELECT price FROM price WHERE id = 1').fetchone()["price"]

    except Exception as e:
        print(e)
        flash("Error al obtener el historial asistido", "error")
        reservations = []
        current_price = 0

    return render_template("admin/assisted_history.html", reservations=reservations, selected_date=selected_date.strftime("%Y-%m-%d"), current_price=current_price)

@bp.route("/finalize-assisted", methods=["POST"])
@auth.only_admin
def finalize_assisted():
    db = get_db()
    reservation_id = request.form.get("reservation_id")
    
    if not reservation_id:
        flash("ID de reservación requerido.", "error")
        return redirect(url_for("admin.assisted_history"))
    
    try:
        reservation = db.execute('SELECT * FROM assisted_reservation WHERE id = ?', (reservation_id,)).fetchone()
        if not reservation or reservation["status"] == "completed":
            flash("Reservación no válida o ya completada.", "error")
            return redirect(url_for("admin.assisted_history"))

        exit_datetime = int(datetime.now(mexico_tz).timestamp())
        entry_datetime = reservation["entry_datetime"]
        
        # Calcular horas (redondeando hacia arriba, mínimo 1 hora)
        hours = ceil((exit_datetime - entry_datetime) / 3600)
        if hours <= 0: 
            hours = 1 
            
        current_price = db.execute('SELECT price FROM price WHERE id = 1').fetchone()["price"]
        total_cost = hours * current_price

        # Actualizar base de datos
        db.execute(
            'UPDATE assisted_reservation SET exit_datetime = ?, status = ?, cost = ? WHERE id = ?',
            (exit_datetime, "completed", total_cost, reservation_id)
        )
        db.commit()
        flash("Reservación finalizada y pagada correctamente.", "success")
    except Exception as e:
        print("Error al finalizar:", e)
        flash("Error al procesar el pago.", "error")
        
    return redirect(url_for("admin.assisted_history"))

@bp.route("/reports")
@auth.only_admin
def reports():
    db = get_db()

    # Conteos filtrados
    app_count = db.execute('SELECT COUNT(*) FROM reservation WHERE status != "canceled" AND type = "hour"').fetchone()[0]
    assisted_count = db.execute('SELECT COUNT(*) FROM assisted_reservation WHERE status != "canceled" AND type = "hour"').fetchone()[0]
    pension_count = db.execute('SELECT COUNT(*) FROM assisted_reservation WHERE status != "canceled" AND type = "pension"').fetchone()[0]

    # Ingresos
    ing_app = db.execute('SELECT SUM(cost) FROM reservation WHERE status = "completed"').fetchone()[0] or 0
    ing_asistido = db.execute('SELECT SUM(cost) FROM assisted_reservation WHERE status = "completed" AND type = "hour"').fetchone()[0] or 0
    ing_pension = db.execute('SELECT SUM(cost) FROM assisted_reservation WHERE status = "completed" AND type = "pension"').fetchone()[0] or 0

    return render_template(
        "admin/reportes.html", 
        app_count=app_count, 
        assisted_count=assisted_count,
        pension_count=pension_count,
        ingresos_app=ing_app,
        ingresos_asistido=ing_asistido,
        ingresos_pension=ing_pension
    )

@bp.route("/reports/export")
@auth.only_admin
def export_report():
    start_date_str = request.args.get('start_date')  # Recibe "YYYY-MM-DD"
    end_date_str = request.args.get('end_date')      # Recibe "YYYY-MM-DD"
    file_format = request.args.get('format')

    db = get_db()

    # CONVERSIÓN DE FECHAS A TIMESTAMPS UNIX (INTEGER)
    date_filter_res = ""
    date_filter_ast = ""
    params = []

    if start_date_str and end_date_str:
        start_timestamp = int(datetime.strptime(f"{start_date_str} 00:00:00", "%Y-%m-%d %H:%M:%S").timestamp())
        end_timestamp = int(datetime.strptime(f"{end_date_str} 23:59:59", "%Y-%m-%d %H:%M:%S").timestamp())
        
        date_filter_res = " AND entry_datetime BETWEEN ? AND ?"
        date_filter_ast = " AND entry_datetime BETWEEN ? AND ?"
        params = [start_timestamp, end_timestamp]

    # 1. Ejecución de Conteos por Canal
    app_count = db.execute(
        f'SELECT COUNT(*) FROM reservation WHERE entry_datetime IS NOT NULL AND type = "hour" {date_filter_res}', 
        params
    ).fetchone()[0]
    
    assisted_count = db.execute(
        f'SELECT COUNT(*) FROM assisted_reservation WHERE status != "canceled" AND type = "hour" {date_filter_ast}', 
        params
    ).fetchone()[0]
    
    pension_count = db.execute(
        f'SELECT COUNT(*) FROM assisted_reservation WHERE status != "canceled" AND type = "pension" {date_filter_ast}', 
        params
    ).fetchone()[0]

    # 2. Sumas de Ingresos Financieros
    ing_app = db.execute(
        f'SELECT SUM(cost) FROM reservation WHERE cost IS NOT NULL {date_filter_res}', 
        params
    ).fetchone()[0] or 0
    
    ing_asistido = db.execute(
        f'SELECT SUM(cost) FROM assisted_reservation WHERE status = "completed" AND type = "hour" {date_filter_ast}', 
        params
    ).fetchone()[0] or 0
    
    ing_pension = db.execute(
        f'SELECT SUM(cost) FROM assisted_reservation WHERE status = "completed" AND type = "pension" {date_filter_ast}', 
        params
    ).fetchone()[0] or 0

    total_ingresos = ing_app + ing_asistido + ing_pension
    total_usuarios = app_count + assisted_count + pension_count

    # ==========================================
    # FLUJO A: GENERACIÓN EXCEL PROFESIONAL (.XLSX)
    # ==========================================
    if file_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen de Ingresos"
        
        # Habilitar líneas de cuadrícula para que no se vea plano
        ws.views.sheetView[0].showGridLines = True
        
        # Definición de Estilos Oficiales de Marca
        font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_sub = Font(name="Segoe UI", size=11, italic=True, color="555555")
        font_data = Font(name="Segoe UI", size=11)
        font_total = Font(name="Segoe UI", size=11, bold=True, color="000000")
        
        fill_brand = PatternFill(start_color="1477D4", end_color="1477D4", fill_type="solid") # Azul corporativo
        fill_header = PatternFill(start_color="1C3455", end_color="1C3455", fill_type="solid") # Azul oscuro
        fill_total = PatternFill(start_color="E9F5FF", end_color="E9F5FF", fill_type="solid") # Azul claro sutil
        
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        # 1. Banner Principal de Título
        ws.merge_cells('A1:C1')
        ws['A1'] = "TIMEPARKING - REPORTE CONSOLIDADO DE INGRESOS"
        ws['A1'].font = font_title
        ws['A1'].fill = fill_brand
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # 2. Rango de Fechas
        ws.merge_cells('A2:C2')
        ws['A2'] = f"Periodo de Auditoría: Desde {start_date_str} Hasta {end_date_str}"
        ws['A2'].font = font_sub
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22
        
        # 3. Cabeceras de la Tabla de Datos
        headers = ["Línea de Negocio / Canal", "Flujo de Vehículos", "Ingresos Netos ($ MXN)"]
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[4].height = 25

        # 4. Inyección de Registros de Datos
        data_rows = [
            ["Aplicación Móvil (Por hora)", app_count, ing_app],
            ["Registro Asistido (Taquilla Caja)", assisted_count, ing_asistido],
            ["Control General de Pensiones", pension_count, ing_pension]
        ]
        
        current_row = 5
        for item in data_rows:
            r_cell = ws.cell(row=current_row, column=1, value=item[0])
            u_cell = ws.cell(row=current_row, column=2, value=item[1])
            i_cell = ws.cell(row=current_row, column=3, value=item[2])
            
            # Formatos y Fuentes
            r_cell.font = font_data
            u_cell.font = font_data
            i_cell.font = font_data
            
            u_cell.number_format = '#,##0'
            i_cell.number_format = '$#,##0.00'
            
            r_cell.alignment = Alignment(horizontal="left", vertical="center")
            u_cell.alignment = Alignment(horizontal="center", vertical="center")
            i_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            r_cell.border = thin_border
            u_cell.border = thin_border
            i_cell.border = thin_border
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
        # 5. Fila de Totales del Sistema
        ws.cell(row=current_row, column=1, value="TOTALIZADORES GENERALES").font = font_total
        ws.cell(row=current_row, column=2, value=total_usuarios).font = font_total
        ws.cell(row=current_row, column=3, value=total_ingresos).font = font_total
        
        ws.cell(row=current_row, column=2).number_format = '#,##0'
        ws.cell(row=current_row, column=3).number_format = '$#,##0.00'
        
        for col_idx in range(1, 4):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = fill_total
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                
        ws.row_dimensions[current_row].height = 24
        
        # 6. Autoajuste de Anchos de Columna para Evitar Texto Cortado
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1 or cell.row == 2:
                    continue # Ignorar celdas combinadas de cabecera
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        # Volcar flujo de bytes en memoria hacia la respuesta HTTP
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        
        response = make_response(excel_stream.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=Reporte_Timeparking_{start_date_str}_a_{end_date_str}.xlsx"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response

    # ==========================================
    # FLUJO B: RENDERIZADO PARA CONVERSIÓN PDF
    # ==========================================
    return render_template(
        "admin/reporte_pdf_template.html",
        start_date=start_date_str,
        end_date=end_date_str,
        app_count=app_count,
        assisted_count=assisted_count,
        pension_count=pension_count,
        ingresos_app=ing_app,
        ingresos_asistido=ing_asistido,
        ingresos_pension=ing_pension,
        total_ingresos=total_ingresos
    )

@bp.route('/register-pension', methods=['POST'])
@auth.only_admin
def register_pension():
    db = get_db()
    first_name = request.form['first_name']
    last_name = request.form['last_name']
    space_id = int(request.form['space_id'])
    start_str = request.form['start_date']
    end_str = request.form['end_date']

    try:
        # Convertir fechas (son solo fecha, sin hora)
        start_dt = datetime.strptime(start_str, "%Y-%m-%d")
        end_dt = datetime.strptime(end_str, "%Y-%m-%d")
        
        # Calcular días (mínimo 1)
        delta = end_dt - start_dt
        days = max(1, delta.days)

        # Obtener precio diario
        price_data = db.execute('SELECT daily_price FROM price WHERE id = 1').fetchone()
        daily_rate = price_data['daily_price'] if price_data else 200.0
        total_cost = days * daily_rate

        # Generar código
        start_ts = int(mexico_tz.localize(start_dt).timestamp())
        end_ts = int(mexico_tz.localize(end_dt).timestamp())
        timestamp_now = int(datetime.now(mexico_tz).timestamp())
        code = f"P-{space_id}-{timestamp_now}"

        db.execute(
            '''INSERT INTO assisted_reservation 
               (first_name, last_name, space_id, entry_datetime, end_datetime, status, cost, code, type)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (first_name, last_name, space_id, start_ts, end_ts, "active", total_cost, code, "pension")
        )
        db.commit()
        flash(f"Pensión registrada para {first_name} por {days} días. Total: ${total_cost}", "success")
    except Exception as e:
        print(e)
        flash("Error al registrar la pensión", "error")

    return redirect(url_for("admin.index"))

@bp.route("/change-price")
@auth.only_admin
def change_price():
    db = get_db()
    
    # Intentamos obtener ambos valores de la URL
    new_hourly = request.args.get("hourlyRate")
    new_daily = request.args.get("dailyRate")
    
    # Si el usuario solo cambió uno, necesitamos mantener el otro como estaba
    # por eso en el HTML pusimos un "hidden" input o simplemente validamos aquí
    if not new_hourly or not new_daily:
        flash("Faltan datos para actualizar los precios.", "error")
        return redirect(url_for("admin.index"))

    try:
        # Esta línea actualiza AMBOS campos en la tabla 'price'
        db.execute(
            'UPDATE price SET price = ?, daily_price = ? WHERE id = 1',
            (new_hourly, new_daily)
        )
        db.commit()
        flash("Tarifas actualizadas correctamente.", "success")
    except Exception as e:
        print(f"Error en DB: {e}")
        flash("No se pudieron guardar los precios.", "error")

    return redirect(url_for("admin.index"))

@bp.route("/pensions")
@auth.only_admin
def pensions():
    db = get_db()
    # Obtenemos todas las pensiones (activas primero)
    raw_pensions = db.execute(
        'SELECT * FROM assisted_reservation WHERE type = "pension" ORDER BY status ASC, entry_datetime DESC'
    ).fetchall()
    
    pensions_list = []
    for p in raw_pensions:
        item = dict(p)
        item["start_str"] = datetime.fromtimestamp(item["entry_datetime"], mexico_tz).strftime("%Y-%m-%d")
        item["end_str"] = datetime.fromtimestamp(item["end_datetime"], mexico_tz).strftime("%Y-%m-%d")
        item["full_name"] = f"{item['first_name']} {item['last_name']}"
        pensions_list.append(item)

    return render_template("admin/pensiones.html", pensions=pensions_list)

@bp.route("/finalize-pension", methods=["POST"])
@auth.only_admin
def finalize_pension():
    db = get_db()
    pension_id = request.form.get("pension_id")
    final_cost = request.form.get("final_cost")

    try:
        # Al marcar como completed, el espacio se libera automáticamente 
        # gracias a la lógica de disponibilidad que ya hicimos.
        db.execute(
            'UPDATE assisted_reservation SET status = "completed", cost = ? WHERE id = ?',
            (final_cost, pension_id)
        )
        db.commit()
        flash("Pensión finalizada y espacio liberado.", "success")
    except:
        flash("Error al finalizar la pensión.", "error")
        
    return redirect(url_for("admin.pensions"))

@bp.route("/manuales")
@auth.only_admin
def manuales():
    return render_template("admin/manuales.html")